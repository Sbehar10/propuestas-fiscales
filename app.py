import streamlit as st
import pandas as pd
import io
import constantes
from datetime import datetime
from constantes import (
    LISTA_PUESTOS, PUESTOS_PROFESIONALES, PRIMA_RIESGO,
    SALARIO_MINIMO_MENSUAL, ISN_TASAS_ESTADO,
)
from motor_calculo import (
    calcular_esquema_actual, calcular_esquema_irt, calcular_excedentes,
    calcular_sociedad_civil, calcular_grupo_nomina, neto_a_bruto
)
from procesador_nomina import (
    detectar_columnas, detectar_bruto_neto, detectar_periodo,
    convertir_a_bruto_mensual, mapear_puestos, validar_datos
)
from generador_word import generar_propuesta_word, fmt_moneda


def fmt_tarjeta(valor):
    """Formato abreviado para tarjetas: $6.04M, $604.5K, $999.00"""
    av = abs(valor)
    signo = "-" if valor < 0 else ""
    if av >= 1_000_000:
        return f"{signo}${av/1_000_000:.2f}M"
    elif av >= 1_000:
        return f"{signo}${av/1_000:.1f}K"
    return f"{signo}${av:,.2f}"


# === CONFIGURACION DE PAGINA ===
st.set_page_config(
    page_title="Sistema de Cotizacion Fiscal 2026",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# === ESTILOS CSS ===
st.markdown("""
<style>
    /* --- Header banner --- */
    .main-header {
        background: linear-gradient(135deg, #1B3A5C 0%, #2C5F8A 100%);
        padding: 2.2rem 2rem;
        border-radius: 8px;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 2px 8px rgba(27,58,92,0.15);
    }
    .main-header h1 {
        color: #FFFFFF !important;
        font-size: 1.8rem !important;
        font-weight: 700 !important;
        margin: 0 !important;
        letter-spacing: 0.3px;
    }
    .main-header .accent {
        color: #C9A962 !important;
    }
    .main-header p {
        color: rgba(255,255,255,0.75) !important;
        font-size: 0.9rem;
        font-weight: 400;
        margin: 0.4rem 0 0 0;
    }

    /* --- Section headers --- */
    .section-header {
        background: transparent;
        padding: 0 0 0.5rem 0;
        border-bottom: 2px solid #E2E8F0;
        margin: 1.8rem 0 1rem 0;
    }
    .section-header h3 {
        color: #1B3A5C !important;
        margin: 0 !important;
        font-size: 1.05rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.2px;
    }

    /* --- Metric cards --- */
    .metric-card {
        background: #FFFFFF;
        padding: 1.4rem 1.2rem;
        border-radius: 8px;
        text-align: center;
        border: 1px solid #E2E8F0;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
        margin-bottom: 0.5rem;
    }
    .metric-card h3 {
        color: #718096 !important;
        font-size: 0.7rem !important;
        font-weight: 600 !important;
        margin: 0 !important;
        text-transform: uppercase;
        letter-spacing: 0.8px;
    }
    .metric-card p {
        color: #1B3A5C !important;
        font-size: 1.7rem !important;
        font-weight: 700 !important;
        margin: 0.4rem 0 0 0 !important;
    }
    .metric-card .sub {
        color: #A0AEC0 !important;
        font-size: 0.72rem !important;
        font-weight: 400 !important;
        margin: 0.15rem 0 0 0 !important;
    }

    /* --- Ahorro cards (soft green) --- */
    .ahorro-verde {
        background: #E8F5E9;
        padding: 1.4rem 1.2rem;
        border-radius: 8px;
        text-align: center;
        border: 1px solid #C8E6C9;
        box-shadow: 0 1px 4px rgba(39,174,96,0.08);
        margin-bottom: 0.5rem;
    }
    .ahorro-verde h3 {
        color: #2E7D32 !important;
        font-size: 0.7rem !important;
        font-weight: 600 !important;
        margin: 0 !important;
        text-transform: uppercase;
        letter-spacing: 0.8px;
    }
    .ahorro-verde p {
        color: #1B5E20 !important;
        font-size: 1.7rem !important;
        font-weight: 700 !important;
        margin: 0.4rem 0 0 0 !important;
    }
    .ahorro-verde .sub {
        color: #4CAF50 !important;
        font-size: 0.72rem !important;
        font-weight: 400 !important;
    }

    /* --- Estado badge --- */
    .estado-badge {
        display: inline-block;
        background: #F7F1E0;
        color: #8B6914 !important;
        padding: 0.25rem 0.7rem;
        border-radius: 4px;
        font-size: 0.78rem;
        font-weight: 600;
        margin-top: 0.3rem;
        border: 1px solid #E8D9A8;
    }

    /* --- Progress bar (cotizador) --- */
    .progress-bar {
        display: flex;
        justify-content: space-between;
        margin: 0.8rem 0 1.5rem 0;
        padding: 0;
        background: #F7FAFC;
        border-radius: 6px;
    }
    .progress-step {
        flex: 1;
        text-align: center;
        padding: 0.55rem 0.3rem;
        font-size: 0.73rem;
        font-weight: 600;
        color: #CBD5E0;
        border-bottom: 3px solid #E2E8F0;
        transition: all 0.2s ease;
    }
    .progress-step.active {
        color: #1B3A5C;
        border-bottom-color: #C9A962;
    }
    .progress-step.done {
        color: #2E7D32;
        border-bottom-color: #27AE60;
    }

    /* --- Buttons --- */
    .stButton>button {
        width: 100%;
    }
    .stButton>button[kind="primary"] {
        background-color: #1B3A5C !important;
        color: #FFFFFF !important;
        border: none !important;
        font-weight: 600 !important;
        border-radius: 6px !important;
    }
    .stButton>button[kind="primary"]:hover {
        background-color: #C9A962 !important;
        color: #1B3A5C !important;
    }

    /* --- Comparison table --- */
    .comp-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.88rem;
        margin: 0.8rem 0;
    }
    .comp-table th {
        background: #1B3A5C;
        color: #FFFFFF;
        padding: 0.6rem 0.8rem;
        text-align: left;
        font-weight: 600;
        font-size: 0.78rem;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .comp-table td {
        padding: 0.5rem 0.8rem;
        border-bottom: 1px solid #E2E8F0;
        color: #2D3748;
    }
    .comp-table tr:nth-child(even) {
        background: #F7FAFC;
    }
    .comp-table .num {
        text-align: right;
        font-variant-numeric: tabular-nums;
    }
    .comp-table .ahorro-row td {
        background: #E8F5E9;
        font-weight: 700;
        color: #1B5E20;
    }
</style>
""", unsafe_allow_html=True)

# === HEADER ===
st.markdown("""
<div class="main-header">
    <h1>Sistema de Cotizacion <span class="accent">Fiscal</span></h1>
    <p>Cotizacion y generacion de propuestas para clientes &mdash; Ano fiscal 2026</p>
</div>
""", unsafe_allow_html=True)


# ============================================================
# HELPERS
# ============================================================

def _set_isn(tasa_isn):
    """Set ISN_TASA globally before calculations (single-threaded Streamlit).
    Must patch both constantes AND motor_calculo since motor_calculo uses 'from constantes import *'
    which copies values at import time."""
    import motor_calculo
    constantes.ISN_TASA = tasa_isn
    motor_calculo.ISN_TASA = tasa_isn


def calcular_base_imss_inteligente(sueldo_bruto, min_profesional, nivel="Moderado"):
    """
    Determina base IMSS segun nivel de sueldo y agresividad.

    Reglas Moderado (default):
    - sueldo <= SM           -> SM (no hay diferencial)
    - sueldo <= min_prof*1.3 -> SM
    - sueldo <= 50,000       -> min_profesional
    - sueldo > 50,000        -> max(min_profesional, sueldo * 0.30)

    Conservador: un nivel arriba (mas base, menos ahorro, menos riesgo)
    Agresivo: un nivel abajo (menos base, mas ahorro, mas riesgo)
    """
    SM = SALARIO_MINIMO_MENSUAL

    if nivel == "Conservador":
        if sueldo_bruto <= SM:
            return SM
        elif sueldo_bruto <= min_profesional * 1.5:
            return min_profesional
        elif sueldo_bruto <= 50_000:
            return min_profesional
        else:
            return max(min_profesional, sueldo_bruto * 0.40)

    elif nivel == "Moderado":
        if sueldo_bruto <= SM:
            return SM
        elif sueldo_bruto <= min_profesional * 1.3:
            return SM
        elif sueldo_bruto <= 50_000:
            return min_profesional
        else:
            return max(min_profesional, sueldo_bruto * 0.30)

    else:  # Agresivo
        if sueldo_bruto <= SM:
            return SM
        elif sueldo_bruto <= min_profesional * 1.1:
            return SM
        elif sueldo_bruto <= 50_000:
            return SM
        else:
            return min_profesional


def generar_excel_resultados(resultados_grupos, nombre_empresa, estado_nombre, tasa_isn):
    """Generate Excel workbook with summary + detail sheets using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # -- Hoja Resumen --
    ws = wb.active
    ws.title = "Resumen"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    gold_font = Font(bold=True, color="C9A962", size=12)
    border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Propuesta Fiscal — {nombre_empresa or 'Cliente'}"
    ws["A1"].font = Font(bold=True, size=14, color="1B3A5C")
    ws["A2"] = f"Estado: {estado_nombre} (ISN {tasa_isn*100:.1f}%)"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A3"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
    ws["A3"].font = Font(size=10, color="666666")

    # Summary table
    row = 5
    headers = ["Puesto", "Empleados", "Sueldo Bruto", "Costo Actual", "Costo IRT", "Ahorro Mensual"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_actual = 0
    total_irt = 0
    total_ahorro = 0

    for r in resultados_grupos:
        row += 1
        ws.cell(row=row, column=1, value=r["puesto"])
        ws.cell(row=row, column=2, value=r["num_empleados"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=round(r["sueldo_bruto"], 2)).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=round(r["actual"]["costo_total"], 2)).number_format = '$#,##0.00'
        costo_irt = r["irt"]["subtotal_factura"] * r["num_empleados"]
        ws.cell(row=row, column=5, value=round(costo_irt, 2)).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=round(r["ahorro_mensual"], 2)).number_format = '$#,##0.00'
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = border

        total_actual += r["actual"]["costo_total"]
        total_irt += costo_irt
        total_ahorro += r["ahorro_mensual"]

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_actual, 2)).number_format = '$#,##0.00'
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(total_irt, 2)).number_format = '$#,##0.00'
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_ahorro, 2)).number_format = '$#,##0.00'
    ws.cell(row=row, column=6).font = gold_font

    # Projection
    row += 2
    ws.cell(row=row, column=1, value="Proyeccion de Ahorro").font = Font(bold=True, size=12)
    row += 1
    for label, mult in [("Mensual", 1), ("Anual", 12), ("2 Anos", 24), ("3 Anos", 36)]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(total_ahorro * mult, 2)).number_format = '$#,##0.00'
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    # -- Hoja Detalle --
    ws2 = wb.create_sheet("Detalle por Grupo")
    detail_row = 1
    for r in resultados_grupos:
        ws2.cell(row=detail_row, column=1, value=r["puesto"]).font = Font(bold=True, size=11)
        ws2.cell(row=detail_row, column=2, value=f"{r['num_empleados']} empleados")
        detail_row += 1

        detail_items = [
            ("Sueldo bruto", r["sueldo_bruto"]),
            ("Base nomina IRT", r["irt"]["base_nomina"]),
            ("Excedente IRT", r["irt"]["excedente_irt"]),
            ("IMSS patronal", r["irt"]["imss_patronal"]["total"]),
            ("INFONAVIT", r["irt"]["infonavit"]),
            ("ISN", r["irt"]["isn"]),
            ("Costo social", r["irt"]["costo_social"]),
            ("Comision", r["irt"]["comision"]),
            ("IVA", r["irt"]["iva"]),
            ("Total factura (por emp)", r["irt"]["total_factura"]),
            ("Costo actual (total grupo)", r["actual"]["costo_total"]),
            ("Ahorro mensual", r["ahorro_mensual"]),
        ]
        for label, val in detail_items:
            ws2.cell(row=detail_row, column=1, value=label)
            ws2.cell(row=detail_row, column=2, value=round(val, 2)).number_format = '$#,##0.00'
            detail_row += 1
        detail_row += 1

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.markdown("### Datos del Cliente")
    nombre_empresa = st.text_input("Nombre de la empresa", placeholder="Grupo Industrial del Norte")
    contacto = st.text_input("Contacto", placeholder="Lic. Roberto Mendez")
    comision_pct = st.slider("Comision (%)", min_value=2.0, max_value=8.0, value=5.0, step=0.5)

    st.markdown("---")

    # Selector de estado para ISN
    st.markdown("### Estado (ISN)")
    estados_lista = sorted(ISN_TASAS_ESTADO.keys())
    idx_cdmx = next((i for i, e in enumerate(estados_lista) if "Ciudad" in e), 0)
    estado_sel = st.selectbox(
        "Estado donde opera la empresa",
        options=estados_lista,
        index=idx_cdmx,
        key="estado_isn"
    )
    tasa_isn = ISN_TASAS_ESTADO[estado_sel]
    st.markdown(
        f'<span class="estado-badge">{estado_sel} — ISN {tasa_isn*100:.2f}%</span>',
        unsafe_allow_html=True,
    )

    st.markdown("---")

    # Prima de riesgo de trabajo
    st.markdown("### Prima de Riesgo")
    modo_prima = st.radio(
        "Selecciona la prima de riesgo:",
        options=["Por clase (I-V)", "Numerica (%)"],
        horizontal=True,
        key="modo_prima",
    )
    if modo_prima == "Por clase (I-V)":
        clase_riesgo_global = st.selectbox(
            "Clase de riesgo IMSS",
            options=["I", "II", "III", "IV", "V"],
            key="clase_riesgo_global"
        )
        from constantes import PRIMA_RIESGO as _PR
        prima_riesgo_global = _PR[clase_riesgo_global]
        st.caption(f"Clase {clase_riesgo_global} = {prima_riesgo_global:.5f}%")
    else:
        prima_riesgo_global = st.number_input(
            "Prima de riesgo (%)",
            min_value=0.0,
            max_value=15.0,
            value=0.54355,
            step=0.01,
            format="%.5f",
            key="prima_numerica",
        )
        clase_riesgo_global = "I"  # fallback label

    st.markdown("---")
    st.markdown("### Nivel de Esquema")
    nivel_esquema = st.radio(
        "Agresividad de la base IMSS:",
        options=["Conservador", "Moderado", "Agresivo"],
        index=1,
        key="nivel_esquema",
        horizontal=True,
    )
    st.caption({
        "Conservador": "Mas base IMSS, menos ahorro, menos riesgo",
        "Moderado": "Balance entre ahorro y seguridad",
        "Agresivo": "Menos base IMSS, mas ahorro, mas riesgo",
    }[nivel_esquema])

    st.markdown("---")
    st.markdown("### Tipo de Servicio")
    tipo_servicio = st.radio(
        "Selecciona el servicio:",
        options=[
            "Cotizador (Subir nomina)",
            "Nomina completa (IRT)",
            "Solo excedentes",
            "Sociedad Civil",
        ],
        index=0
    )

    tipo_map = {
        "Cotizador (Subir nomina)": "cotizador",
        "Nomina completa (IRT)": "nomina",
        "Solo excedentes": "excedentes",
        "Sociedad Civil": "sc",
    }
    tipo = tipo_map[tipo_servicio]

# Set ISN for all calculations
_set_isn(tasa_isn)


# ============================================================
# HELPER: TABLA COMPARATIVA HTML
# ============================================================
def tabla_comparativa_irt(r):
    """Genera tabla HTML comparativa Esquema Actual vs IRT para un grupo."""
    act = r["actual"]
    irt = r["irt"]
    n = r["num_empleados"]
    neto_orig = r.get("sueldo_neto_original")

    rows = []
    if neto_orig:
        rows.append(("Sueldo neto del empleado", neto_orig, neto_orig))
        rows.append(("Bruto equivalente", act["sueldo_bruto"], act["sueldo_bruto"]))
    else:
        rows.append(("Sueldo bruto mensual", act["sueldo_bruto"], act["sueldo_bruto"]))
    rows += [
        ("Base nomina (IMSS)", act["sueldo_bruto"], irt["base_nomina"]),
        ("ISR Art. 96", act["isr"]["isr_neto"], irt["isr"]["isr_neto"]),
        ("IMSS patronal", act["imss_patronal"]["total"], irt["imss_patronal"]["total"]),
        ("Infonavit", act["infonavit"], irt["infonavit"]),
        ("ISN", act["isn"], irt["isn"]),
        ("Prestaciones de ley", act["prestaciones"]["total_mensual"], irt["prestaciones"]["total_mensual"]),
        ("Excedente IRT / PPS", 0, irt["excedente_irt"]),
        ("Costo social total", 0, irt["costo_social"]),
        ("Comision", 0, irt["comision"]),
    ]
    html = '<table class="comp-table"><tr><th>Concepto</th><th class="num">Esquema Actual</th><th class="num">Esquema IRT</th><th class="num">Diferencia</th></tr>'
    for label, val_act, val_irt in rows:
        diff = val_act - val_irt
        html += f'<tr><td>{label}</td><td class="num">{fmt_moneda(val_act)}</td><td class="num">{fmt_moneda(val_irt)}</td><td class="num">{fmt_moneda(diff)}</td></tr>'

    # Totals
    costo_act = act["costo_por_empleado"]
    costo_irt = irt["subtotal_factura"]
    ahorro = costo_act - costo_irt
    html += f'<tr class="ahorro-row"><td><strong>Costo empresa (pre-IVA) x1</strong></td><td class="num">{fmt_moneda(costo_act)}</td><td class="num">{fmt_moneda(costo_irt)}</td><td class="num">{fmt_moneda(ahorro)}</td></tr>'
    if n > 1:
        html += f'<tr class="ahorro-row"><td><strong>Total x{n} empleados</strong></td><td class="num">{fmt_moneda(costo_act * n)}</td><td class="num">{fmt_moneda(costo_irt * n)}</td><td class="num">{fmt_moneda(ahorro * n)}</td></tr>'

    # IVA (acreditable — no es costo)
    html += f'<tr><td>IVA (acreditable)</td><td class="num">—</td><td class="num">{fmt_moneda(irt["iva"])}</td><td class="num" colspan="1">—</td></tr>'
    # Neto trabajador
    html += f'<tr><td>Neto trabajador</td><td class="num">{fmt_moneda(act["neto_trabajador"])}</td><td class="num">{fmt_moneda(irt["neto_trabajador"])}</td><td class="num">{fmt_moneda(irt["neto_trabajador"] - act["neto_trabajador"])}</td></tr>'
    html += '</table>'
    html += '<p style="font-size:0.85em;color:#555;margin-top:0.5em;">La comision es 100% deducible para ISR empresarial y el IVA es acreditable.</p>'
    return html


# ============================================================
# HELPER: MOSTRAR RESULTADOS DE NOMINA IRT
# ============================================================
def mostrar_resultados_nomina(resultados_grupos, comision_pct, nombre_empresa, contacto, es_neto=False):
    """Muestra metricas, proyeccion, grafica, detalle y descargas Word+Excel para resultados IRT."""
    import altair as alt

    costo_actual_total = 0
    ahorro_total = 0
    total_empleados = 0
    total_administrado = 0
    total_iva = 0

    for r in resultados_grupos:
        costo_actual_total += r["actual"]["costo_total"]
        ahorro_total += r["ahorro_mensual"]
        total_empleados += r["num_empleados"]
        total_administrado += r["irt"]["total_administrado"] * r["num_empleados"]
        total_iva += r["irt"]["iva"] * r["num_empleados"]

    costo_propuesto_pre_iva = costo_actual_total - ahorro_total
    pct_ahorro = (ahorro_total / costo_actual_total * 100) if costo_actual_total > 0 else 0

    # Mensaje neto cuando aplica
    if es_neto:
        st.success("Sus empleados siguen recibiendo el mismo neto. Lo que cambia es el costo patronal.")

    # Metricas principales
    st.markdown('<div class="section-header"><h3>Resultados del Analisis</h3></div>', unsafe_allow_html=True)

    if es_neto:
        col1, col2, col3, col4, col5 = st.columns(5)
    else:
        col1, col2, col3, col4 = st.columns(4)
        col5 = None
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h3>Costo Actual Mensual</h3>
            <p>{fmt_tarjeta(costo_actual_total)}</p>
            <p class="sub">{fmt_moneda(costo_actual_total)}</p>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <h3>Costo IRT Propuesto</h3>
            <p>{fmt_tarjeta(costo_propuesto_pre_iva)}</p>
            <p class="sub">{fmt_moneda(costo_propuesto_pre_iva)} (pre-IVA)</p>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <h3>Empleados</h3>
            <p>{total_empleados}</p>
            <p class="sub">{len(resultados_grupos)} grupo(s)</p>
        </div>""", unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div class="ahorro-verde">
            <h3>Ahorro Mensual</h3>
            <p>{fmt_tarjeta(ahorro_total)}</p>
            <p class="sub">{fmt_moneda(ahorro_total)} ({pct_ahorro:.1f}%)</p>
        </div>""", unsafe_allow_html=True)
    if col5:
        with col5:
            st.markdown(f"""
            <div class="metric-card">
                <h3>Neto Empleado</h3>
                <p>Sin cambio</p>
                <p class="sub">Mismo deposito</p>
            </div>""", unsafe_allow_html=True)

    # IVA acreditable + nota deducibilidad
    st.markdown(f"**IVA total (acreditable):** {fmt_moneda(total_iva)} — no es costo para la empresa.")
    st.caption("La comision es 100% deducible para ISR empresarial y el IVA es acreditable.")

    # Ahorro validation
    if ahorro_total <= 0:
        st.warning("El ahorro total es negativo o cero. Recomendacion: ajustar las bases IMSS o considerar otro esquema.")
    elif pct_ahorro < 3:
        st.warning(f"El ahorro total es minimo ({pct_ahorro:.1f}%). Recomendacion: reducir bases IMSS o considerar otro esquema para algunos grupos.")

    st.markdown("")

    # Proyeccion de Ahorro
    st.markdown('<div class="section-header"><h3>Proyeccion de Ahorro</h3></div>', unsafe_allow_html=True)
    df_ahorro = pd.DataFrame({
        "Periodo": ["Mensual", "Anual", "2 Anos", "3 Anos"],
        "Ahorro IRT": [fmt_moneda(ahorro_total * m) for m in [1, 12, 24, 36]],
        "% Ahorro": [f"{pct_ahorro:.1f}%"] * 4,
    })
    st.dataframe(df_ahorro, use_container_width=True, hide_index=True)

    # Grafica comparativa
    st.markdown('<div class="section-header"><h3>Comparativo de Costos</h3></div>', unsafe_allow_html=True)
    chart_data = pd.DataFrame({
        "Esquema": ["Actual (100%)", "IRT Propuesto"],
        "Costo Mensual": [costo_actual_total, costo_actual_total - ahorro_total]
    })
    chart = alt.Chart(chart_data).mark_bar(cornerRadiusTopLeft=6, cornerRadiusTopRight=6).encode(
        x=alt.X("Esquema:N", sort=None, axis=alt.Axis(labelAngle=0)),
        y=alt.Y("Costo Mensual:Q", axis=alt.Axis(format="$,.0f")),
        color=alt.Color("Esquema:N", scale=alt.Scale(
            domain=["Actual (100%)", "IRT Propuesto"],
            range=["#1B3A5C", "#27AE60"]
        ), legend=None),
        tooltip=[
            alt.Tooltip("Esquema:N"),
            alt.Tooltip("Costo Mensual:Q", format="$,.2f")
        ]
    ).properties(height=350)
    st.altair_chart(chart, use_container_width=True)

    # Detalle por grupo
    st.markdown('<div class="section-header"><h3>Detalle por Grupo</h3></div>', unsafe_allow_html=True)
    for r in resultados_grupos:
        neto_orig = r.get("sueldo_neto_original")
        if neto_orig:
            label_exp = f"{r['puesto']} — {r['num_empleados']} emp. — Neto {fmt_moneda(neto_orig)} (Bruto {fmt_moneda(r['sueldo_bruto'])})"
        else:
            label_exp = f"{r['puesto']} — {r['num_empleados']} emp. a {fmt_moneda(r['sueldo_bruto'])}"
        with st.expander(label_exp):
            if neto_orig:
                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h3>Neto Empleado</h3>
                        <p>{fmt_tarjeta(neto_orig)}</p>
                        <p class="sub">{fmt_moneda(neto_orig)} — sin cambio</p>
                    </div>""", unsafe_allow_html=True)
                with c2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h3>Costo Actual / mes</h3>
                        <p>{fmt_tarjeta(r["actual"]["costo_total"])}</p>
                        <p class="sub">{fmt_moneda(r["actual"]["costo_total"])}</p>
                    </div>""", unsafe_allow_html=True)
                with c3:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h3>Base Nomina IRT</h3>
                        <p>{fmt_tarjeta(r["irt"]["base_nomina"])}</p>
                        <p class="sub">{fmt_moneda(r["irt"]["base_nomina"])}</p>
                    </div>""", unsafe_allow_html=True)
                with c4:
                    st.markdown(f"""
                    <div class="ahorro-verde">
                        <h3>Ahorro / mes</h3>
                        <p>{fmt_tarjeta(r["ahorro_mensual"])}</p>
                        <p class="sub">{fmt_moneda(r["ahorro_mensual"])}</p>
                    </div>""", unsafe_allow_html=True)
            else:
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h3>Costo Actual / mes</h3>
                        <p>{fmt_tarjeta(r["actual"]["costo_total"])}</p>
                        <p class="sub">{fmt_moneda(r["actual"]["costo_total"])}</p>
                    </div>""", unsafe_allow_html=True)
                with c2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h3>Base Nomina IRT</h3>
                        <p>{fmt_tarjeta(r["irt"]["base_nomina"])}</p>
                        <p class="sub">{fmt_moneda(r["irt"]["base_nomina"])}</p>
                    </div>""", unsafe_allow_html=True)
                with c3:
                    st.markdown(f"""
                    <div class="ahorro-verde">
                        <h3>Ahorro / mes</h3>
                        <p>{fmt_tarjeta(r["ahorro_mensual"])}</p>
                        <p class="sub">{fmt_moneda(r["ahorro_mensual"])}</p>
                    </div>""", unsafe_allow_html=True)

            # Per-group ahorro validation
            pct_ahorro_grupo = (r["ahorro_mensual"] / r["actual"]["costo_total"] * 100) if r["actual"]["costo_total"] > 0 else 0
            if r["ahorro_mensual"] <= 0:
                st.warning(f"Este grupo genera ahorro negativo. Recomendacion: no migrar a IRT o reducir la base IMSS.")
            elif pct_ahorro_grupo < 3:
                st.warning(f"Ahorro minimo ({pct_ahorro_grupo:.1f}%). Considerar reducir base IMSS o usar otro esquema.")

            # Comparison table
            st.markdown(tabla_comparativa_irt(r), unsafe_allow_html=True)

    # === DESCARGAS ===
    st.markdown("---")
    st.markdown('<div class="section-header"><h3>Descargar Propuesta</h3></div>', unsafe_allow_html=True)

    datos_cliente = {
        "nombre_empresa": nombre_empresa or "Cliente",
        "contacto": contacto or "Sin especificar",
        "comision_pct": comision_pct,
    }
    resultados_word = {
        "total_empleados": total_empleados,
        "costo_actual_total": costo_actual_total,
        "ahorro_total_mensual": ahorro_total,
        "total_administrado": total_administrado,
        "es_neto": es_neto,
    }

    buffer_word = generar_propuesta_word(
        datos_cliente=datos_cliente,
        tipo_servicio="nomina",
        resultados=resultados_word,
        grupos=resultados_grupos,
    )
    buffer_excel = generar_excel_resultados(resultados_grupos, nombre_empresa, estado_sel, tasa_isn)

    nombre_base = f"{nombre_empresa or 'Cliente'}_{datetime.now().strftime('%Y%m%d')}"

    col_w, col_e = st.columns(2)
    with col_w:
        st.download_button(
            label="DESCARGAR WORD",
            data=buffer_word,
            file_name=f"Propuesta_{nombre_base}.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            type="primary",
            use_container_width=True,
        )
    with col_e:
        st.download_button(
            label="DESCARGAR EXCEL",
            data=buffer_excel,
            file_name=f"Propuesta_{nombre_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True,
        )


# ============================================================
# COTIZADOR (SUBIR NOMINA)
# ============================================================
if tipo == "cotizador":
    st.markdown('<div class="section-header"><h3>Cotizador — Subir Nomina</h3></div>', unsafe_allow_html=True)
    st.markdown("Sube un archivo Excel o CSV con la nomina del cliente para generar la propuesta automaticamente.")

    # Progress bar
    step = 1
    archivo = st.file_uploader(
        "Sube el archivo de nomina",
        type=["xlsx", "xls", "csv"],
        key="archivo_nomina"
    )
    if archivo:
        step = 2

    steps_html = ""
    step_labels = ["1. Subir archivo", "2. Mapear columnas", "3. Calcular", "4. Resultados"]
    for i, lbl in enumerate(step_labels, 1):
        cls = "done" if i < step else ("active" if i == step else "")
        steps_html += f'<div class="progress-step {cls}">{lbl}</div>'
    st.markdown(f'<div class="progress-bar">{steps_html}</div>', unsafe_allow_html=True)

    if archivo is not None:
        # Leer archivo — smart header detection (prueba header=0..3)
        def _score_header(df_candidate):
            """Score a DataFrame by how many columns have recognizable text names."""
            score = 0
            for c in df_candidate.columns:
                s = str(c).strip()
                if s and s.lower() != "nan" and not s.startswith("Unnamed"):
                    try:
                        float(s)
                    except ValueError:
                        score += 1  # Non-numeric, non-empty name
            return score

        def _fix_col_names(df_in):
            """Fix empty/NaN column names and deduplicate."""
            new_cols = []
            for i, c in enumerate(df_in.columns):
                if c is None or str(c).strip() == "" or str(c).lower() == "nan":
                    new_cols.append(f"Col_{i+1}")
                else:
                    new_cols.append(str(c).strip())
            seen = {}
            deduped = []
            for c in new_cols:
                if c in seen:
                    seen[c] += 1
                    deduped.append(f"{c}_{seen[c]}")
                else:
                    seen[c] = 0
                    deduped.append(c)
            df_in.columns = deduped
            return df_in

        try:
            if archivo.name.endswith(".csv"):
                best_df = None
                best_score = -1
                for h in range(4):
                    try:
                        archivo.seek(0)
                        candidate = pd.read_csv(archivo, encoding="utf-8", header=h)
                        candidate = candidate.dropna(how="all").reset_index(drop=True)
                        s = _score_header(candidate)
                        if s > best_score:
                            best_score = s
                            best_df = candidate
                    except Exception:
                        pass
                if best_df is None:
                    try:
                        archivo.seek(0)
                        best_df = pd.read_csv(archivo, encoding="latin-1")
                    except Exception:
                        archivo.seek(0)
                        best_df = pd.read_csv(archivo)
                df_raw = best_df
            else:
                # Multi-sheet: let user pick sheet
                archivo.seek(0)
                xls = pd.ExcelFile(archivo)
                sheet_names = xls.sheet_names
                if len(sheet_names) > 1:
                    sheet_sel = st.selectbox(
                        "El archivo tiene varias hojas. Selecciona:",
                        options=sheet_names,
                        key="sheet_selector",
                    )
                else:
                    sheet_sel = sheet_names[0]

                # Try header=0..3, pick the one with most recognizable column names
                best_df = None
                best_score = -1
                for h in range(4):
                    try:
                        candidate = pd.read_excel(xls, sheet_name=sheet_sel, header=h)
                        candidate = candidate.dropna(how="all").reset_index(drop=True)
                        s = _score_header(candidate)
                        if s > best_score:
                            best_score = s
                            best_df = candidate
                    except Exception:
                        pass
                if best_df is None:
                    best_df = pd.read_excel(xls, sheet_name=sheet_sel)
                df_raw = best_df

            df_raw = df_raw.dropna(how="all").reset_index(drop=True)
            df_raw = _fix_col_names(df_raw)

        except Exception as e:
            st.error(f"Error al leer el archivo: {e}")
            st.stop()

        st.markdown("#### Vista previa del archivo")
        try:
            st.dataframe(df_raw.head(10), use_container_width=True, hide_index=True)
        except Exception:
            st.dataframe(df_raw.head(10).astype(str), use_container_width=True, hide_index=True)
        st.caption(f"{len(df_raw)} filas x {len(df_raw.columns)} columnas")

        # --- Helper: safe column access (handles duplicate col names) ---
        def _safe_series(df, col_name):
            """Always return a Series, even if duplicate columns yield a DataFrame."""
            result = df[col_name]
            if isinstance(result, pd.DataFrame):
                return result.iloc[:, 0]
            return result

        # --- Auto-detect columns ---
        cols_detectadas = detectar_columnas(df_raw)
        periodo = detectar_periodo(df_raw)

        col_puesto = cols_detectadas["puesto"]
        col_sueldo = cols_detectadas["sueldo"]
        col_empleados = cols_detectadas["num_empleados"]

        if col_sueldo is None:
            st.error("No se pudo detectar la columna de sueldo. Verifica que tu archivo tenga una columna con montos de nómina.")
            st.stop()
        if col_puesto is None:
            st.error("No se pudo detectar la columna de puesto. Verifica que tu archivo tenga una columna con nombres de puestos.")
            st.stop()

        cada_fila_un_empleado = col_empleados is None

        _msg_cols = f"Puesto → **{col_puesto}** | Sueldo → **{col_sueldo}**"
        if not cada_fila_un_empleado:
            _msg_cols += f" | Empleados → **{col_empleados}**"
        st.success(f"✅ Columnas detectadas automáticamente: {_msg_cols}")

        # --- Periodo de nomina (auto-detect sets default) ---
        st.markdown("#### Periodo de la nomina")
        _periodo_cap = {"quincenal": "Quincenal", "semanal": "Semanal", "mensual": "Mensual"}.get(periodo, "Mensual")
        st.session_state["periodo_nomina"] = _periodo_cap
        periodo_nomina = st.radio(
            "Los sueldos del archivo son:",
            options=["Mensual", "Quincenal", "Catorcenal", "Semanal"],
            key="periodo_nomina",
            horizontal=True,
        )
        multiplicadores_periodo = {
            "Mensual": 1.0,
            "Quincenal": 2.0,
            "Catorcenal": 365.0 / 14.0 / 12.0,  # 2.1726
            "Semanal": 365.0 / 7.0 / 12.0,       # 4.3452
        }
        mult_periodo = multiplicadores_periodo[periodo_nomina]
        if periodo_nomina != "Mensual":
            st.info(f"Los sueldos se multiplicaran por **{mult_periodo:.4f}** para convertir a mensual.")

        columnas_df = list(df_raw.columns)

        # --- Columnas adicionales de ingreso (opcional) ---
        st.markdown("#### Columnas adicionales de ingreso (opcional)")
        st.caption("Si el archivo tiene otras columnas de pago ademas del sueldo base, seleccionalas para capturar el ingreso total real.")

        # Exclude already-mapped columns from options
        cols_excluir = {col_puesto, col_sueldo}
        if not cada_fila_un_empleado:
            cols_excluir.add(col_empleados)
        cols_adicionales_opciones = [c for c in columnas_df if c not in cols_excluir]

        cols_adicionales_sel = st.multiselect(
            "Columnas adicionales de ingreso",
            options=cols_adicionales_opciones,
            default=[],
            key="cols_ingreso_adicional",
        )

        TIPOS_COLUMNA_INGRESO = [
            "Nomina formal (con IMSS)",
            "Ingreso exento adicional",
            "Otro ingreso",
        ]
        clasificacion_cols_adicionales = {}
        if cols_adicionales_sel:
            st.caption("Clasifica cada columna — solo 'Nomina formal' genera cargas sociales patronales:")
            for i, col_ad in enumerate(cols_adicionales_sel):
                clasificacion_cols_adicionales[col_ad] = st.selectbox(
                    f"Tipo: {col_ad}",
                    options=TIPOS_COLUMNA_INGRESO,
                    index=1,
                    key=f"tipo_col_ad_{i}",
                )

        tiene_cols_adicionales = len(cols_adicionales_sel) > 0

        # --- Bruto / Neto ---
        st.markdown("#### Tipo de sueldo")
        tipo_sueldo_detectado = detectar_bruto_neto(df_raw, col_sueldo)

        col1, col2 = st.columns(2)
        with col1:
            if tipo_sueldo_detectado != "desconocido":
                st.session_state["tipo_sueldo"] = tipo_sueldo_detectado.capitalize()
            tipo_sueldo = st.radio(
                "El sueldo del archivo es:",
                options=["Bruto", "Neto"],
                key="tipo_sueldo",
                horizontal=True,
            )
            if tipo_sueldo_detectado != "desconocido":
                st.caption(f"Auto-detectado: **{tipo_sueldo_detectado}**")
        with col2:
            st.markdown(f"**Prima de riesgo:** {prima_riesgo_global:.5f}%")
            st.caption("(Configurada en el panel lateral)")

        # --- Puesto mapping ---
        st.markdown("#### Mapeo de puestos al catalogo")

        try:
            cols_trabajo = [col_puesto, col_sueldo]
            if not cada_fila_un_empleado:
                cols_trabajo.append(col_empleados)
            # Verify all columns exist
            for c in cols_trabajo:
                if c not in df_raw.columns:
                    st.error(f"La columna '{c}' no se encontro en el archivo.")
                    st.stop()

            df_trabajo = df_raw[cols_trabajo].copy()
            df_trabajo[col_sueldo] = pd.to_numeric(_safe_series(df_trabajo, col_sueldo), errors="coerce")
            df_trabajo = df_trabajo.dropna(subset=[col_sueldo])

            if not cada_fila_un_empleado:
                df_trabajo[col_empleados] = pd.to_numeric(_safe_series(df_trabajo, col_empleados), errors="coerce").fillna(1).astype(int)

            # --- ERROR 1 FIX: Filtrar filas de totales/sumas ---
            filas_antes = len(df_trabajo)
            # Remove rows where puesto column says TOTAL/SUMA/SUBTOTAL or is empty
            puesto_series = _safe_series(df_trabajo, col_puesto).astype(str).str.strip().str.upper()
            mask_total = puesto_series.isin(["TOTAL", "SUMA", "SUBTOTAL", "GRAN TOTAL", "NAN", ""])
            # Remove rows where all text columns are empty but sueldo has a number
            mask_empty_text = puesto_series.isin(["NAN", "", "NONE"])
            # Remove row if its sueldo equals the sum of all other sueldos (likely a total row)
            sueldo_vals = _safe_series(df_trabajo, col_sueldo)
            suma_todos = sueldo_vals.sum()
            mask_es_suma = False
            if len(sueldo_vals) > 2:
                for idx in sueldo_vals.index:
                    val = sueldo_vals.loc[idx]
                    resto = suma_todos - val
                    if abs(val - resto) < 1.0 and val > 0:
                        mask_total = mask_total | (sueldo_vals.index == idx)
            df_trabajo = df_trabajo[~(mask_total | mask_empty_text)].reset_index(drop=True)
            filas_eliminadas = filas_antes - len(df_trabajo)
            if filas_eliminadas > 0:
                st.info(f"Se eliminaron **{filas_eliminadas}** fila(s) de totales/sumas detectadas.")

            if len(df_trabajo) == 0:
                st.error("No se encontraron filas con valores numericos validos en la columna de sueldo.")
                st.stop()

            df_mapeo = mapear_puestos(_safe_series(df_trabajo, col_puesto))

            expected_cols = ["puesto_original", "puesto_catalogo", "minimo_profesional", "confianza"]
            missing_cols = [c for c in expected_cols if c not in df_mapeo.columns]
            if missing_cols:
                st.error(f"Error interno en mapeo de puestos: columnas faltantes {missing_cols}")
                st.stop()

        except Exception as e:
            st.error(f"Error al procesar las columnas: {e}")
            st.stop()

        opciones_catalogo = sorted([p for p in PUESTOS_PROFESIONALES.keys()])

        df_editor = df_mapeo[["puesto_original", "puesto_catalogo", "minimo_profesional", "confianza"]].copy()
        df_editor.columns = ["Puesto Original", "Puesto Catalogo", "Min. Profesional", "Confianza"]

        baja_confianza = (df_editor["Confianza"] < 0.7).sum()
        if baja_confianza > 0:
            st.warning(f"Se encontraron **{baja_confianza}** puestos con baja confianza de mapeo. Revisa y corrige si es necesario.")

        df_editado = st.data_editor(
            df_editor,
            column_config={
                "Puesto Original": st.column_config.TextColumn("Puesto Original", disabled=True),
                "Puesto Catalogo": st.column_config.SelectboxColumn(
                    "Puesto Catalogo",
                    options=opciones_catalogo,
                    required=True,
                ),
                "Min. Profesional": st.column_config.NumberColumn("Min. Profesional", format="$%d"),
                "Confianza": st.column_config.ProgressColumn("Confianza", min_value=0, max_value=1, format="%.0f%%"),
            },
            use_container_width=True,
            hide_index=True,
            key="editor_puestos",
        )

        mapeo_final = {}
        for _, row in df_editado.iterrows():
            puesto_cat = row["Puesto Catalogo"]
            # Use the edited Min. Profesional value (user may have changed it)
            min_prof = row["Min. Profesional"]
            if pd.isna(min_prof) or min_prof <= 0:
                min_prof = PUESTOS_PROFESIONALES.get(puesto_cat, SALARIO_MINIMO_MENSUAL)
            mapeo_final[row["Puesto Original"]] = {
                "puesto_catalogo": puesto_cat,
                "minimo_profesional": min_prof,
            }

        # --- Validacion ---
        cols_val = {"puesto": col_puesto, "sueldo": col_sueldo,
                    "num_empleados": col_empleados if not cada_fila_un_empleado else None}
        warnings = validar_datos(df_raw, cols_val)
        if warnings:
            for w in warnings:
                st.warning(w)

        # --- Calcular ---
        st.markdown("---")
        if st.button("CALCULAR PROPUESTA", type="primary", use_container_width=True, key="calc_cotizador"):
            with st.spinner("Calculando propuesta..."):
                # Sumar columnas adicionales de ingreso al sueldo si aplica
                # "Nomina formal" → se suma al sueldo bruto (genera cargas patronales)
                # "Ingreso exento" / "Otro ingreso" → se suma solo al costo (sin cargas)
                if tiene_cols_adicionales and cols_adicionales_sel:
                    try:
                        cols_formal = []
                        cols_exento = []
                        for col_ad in cols_adicionales_sel:
                            if col_ad not in df_raw.columns:
                                continue
                            # Extract values aligned to df_trabajo index
                            raw_vals = df_raw.loc[df_trabajo.index, col_ad]
                            if isinstance(raw_vals, pd.DataFrame):
                                raw_vals = raw_vals.iloc[:, 0]
                            df_trabajo[f"_ing_{col_ad}"] = pd.to_numeric(
                                raw_vals, errors="coerce"
                            ).fillna(0).values
                            tipo_col = clasificacion_cols_adicionales.get(col_ad, "Otro ingreso")
                            if tipo_col == "Nomina formal (con IMSS)":
                                cols_formal.append(f"_ing_{col_ad}")
                            else:
                                cols_exento.append(f"_ing_{col_ad}")

                        # Nomina formal: suma al sueldo bruto (motor calcula cargas sobre total)
                        if cols_formal:
                            df_trabajo[col_sueldo] = df_trabajo[col_sueldo] + df_trabajo[cols_formal].sum(axis=1)
                            st.info(f"Se sumaron **{len(cols_formal)}** columna(s) de nomina formal al sueldo base.")

                        # Exento/Otro: guardar para sumar al costo total despues
                        if cols_exento:
                            df_trabajo["_ingreso_exento"] = df_trabajo[cols_exento].sum(axis=1)
                            st.info(f"Se detectaron **{len(cols_exento)}** columna(s) de ingreso sin cargas patronales.")
                        else:
                            df_trabajo["_ingreso_exento"] = 0
                    except Exception as e:
                        st.warning(f"No se pudieron procesar las columnas adicionales de ingreso: {e}")
                        df_trabajo["_ingreso_exento"] = 0

                # Ensure _ingreso_exento column exists
                if "_ingreso_exento" not in df_trabajo.columns:
                    df_trabajo["_ingreso_exento"] = 0

                if cada_fila_un_empleado:
                    # Agrupar por puesto catálogo mapeado (no por fila individual)
                    df_trabajo["_puesto_cat"] = _safe_series(df_trabajo, col_puesto).map(
                        lambda p: mapeo_final.get(p, {}).get("puesto_catalogo", "Otro (personalizado)")
                    )
                    df_agrupado = df_trabajo.groupby("_puesto_cat").agg(
                        sueldo_promedio=(col_sueldo, "mean"),
                        num_empleados=(col_sueldo, "count"),
                        exento_promedio=("_ingreso_exento", "mean"),
                    ).reset_index()
                    df_agrupado.rename(columns={"_puesto_cat": col_puesto, "sueldo_promedio": col_sueldo}, inplace=True)
                else:
                    df_agrupado = df_trabajo.rename(columns={col_empleados: "num_empleados"})
                    if "_ingreso_exento" in df_agrupado.columns:
                        df_agrupado.rename(columns={"_ingreso_exento": "exento_promedio"}, inplace=True)
                    else:
                        df_agrupado["exento_promedio"] = 0

                es_neto = tipo_sueldo == "Neto"

                resultados_grupos = []
                conversiones_neto = []  # Track neto→bruto conversions
                ajustes_base = []  # Track auto-adjusted IMSS bases
                for _, fila in df_agrupado.iterrows():
                    puesto_orig = fila[col_puesto]
                    sueldo_raw = fila[col_sueldo]
                    n_emp = int(fila["num_empleados"])

                    if n_emp <= 0:
                        continue

                    # Convert to monthly bruto using auto-detected period and salary type
                    sueldo_bruto = convertir_a_bruto_mensual(sueldo_raw, periodo, tipo_sueldo.lower())
                    sueldo = sueldo_raw * mult_periodo  # Keep for display

                    if es_neto:
                        conversiones_neto.append({"puesto": puesto_orig, "periodo": fmt_moneda(sueldo_raw), "mensual": fmt_moneda(sueldo), "bruto": fmt_moneda(sueldo_bruto), "emp": n_emp})

                    info_puesto = mapeo_final.get(puesto_orig, {
                        "puesto_catalogo": "Otro (personalizado)",
                        "minimo_profesional": SALARIO_MINIMO_MENSUAL,
                    })

                    # Base IMSS inteligente según nivel de esquema
                    min_prof = info_puesto["minimo_profesional"]
                    base_imss = calcular_base_imss_inteligente(sueldo_bruto, min_prof, nivel_esquema)
                    if base_imss != min_prof:
                        info_puesto = dict(info_puesto)  # copy to avoid mutating mapeo_final
                        info_puesto["minimo_profesional"] = base_imss
                        ajustes_base.append({
                            "puesto": puesto_orig,
                            "sueldo": fmt_moneda(sueldo_bruto),
                            "min_prof_original": fmt_moneda(min_prof),
                            "base_ajustada": fmt_moneda(base_imss),
                        })

                    r = calcular_grupo_nomina(
                        puesto=info_puesto["puesto_catalogo"],
                        num_empleados=n_emp,
                        sueldo_bruto=sueldo_bruto,
                        clase_riesgo=clase_riesgo_global,
                        minimo_profesional=info_puesto["minimo_profesional"],
                        comision_pct=comision_pct,
                        prima_riesgo=prima_riesgo_global,
                    )
                    # Store neto original when applicable
                    if es_neto:
                        r["sueldo_neto_original"] = sueldo

                    # Add exento income to actual cost (no cargas, just cost)
                    exento_por_emp = fila.get("exento_promedio", 0)
                    if exento_por_emp and exento_por_emp > 0:
                        exento_mensual = exento_por_emp * mult_periodo
                        r["actual"]["costo_total"] += exento_mensual * n_emp
                        r["actual"]["costo_por_empleado"] += exento_mensual
                        r["sueldo_bruto"] += exento_mensual
                        r["ingreso_exento_adicional"] = exento_mensual
                        # Recalcular ahorro
                        r["ahorro_mensual"] = r["actual"]["costo_total"] - (r["irt"]["subtotal_factura"] * n_emp)
                        r["ahorro_anual"] = r["ahorro_mensual"] * 12

                    resultados_grupos.append(r)

                if resultados_grupos:
                    # Show period conversion info
                    if periodo_nomina != "Mensual":
                        st.markdown(f'<div class="section-header"><h3>Conversion {periodo_nomina} → Mensual (x{mult_periodo:.4f})</h3></div>', unsafe_allow_html=True)

                    # Show neto→bruto conversion table if applicable
                    if conversiones_neto:
                        st.markdown('<div class="section-header"><h3>Conversion Neto a Bruto</h3></div>', unsafe_allow_html=True)
                        df_conv = pd.DataFrame(conversiones_neto)
                        df_conv.columns = ["Puesto", "Sueldo Periodo", "Mensual Estimado", "Bruto Estimado", "Empleados"]
                        st.dataframe(df_conv, use_container_width=True, hide_index=True)

                    # Show auto-adjusted base IMSS warnings
                    if ajustes_base:
                        st.info(f"Se ajusto la base IMSS de **{len(ajustes_base)}** grupo(s) segun nivel **{nivel_esquema}**.")
                        df_ajustes = pd.DataFrame(ajustes_base)
                        df_ajustes.columns = ["Puesto", "Sueldo Bruto", "Min. Profesional", "Base IMSS Ajustada"]
                        st.dataframe(df_ajustes, use_container_width=True, hide_index=True)

                    mostrar_resultados_nomina(resultados_grupos, comision_pct, nombre_empresa, contacto, es_neto=es_neto)
                else:
                    st.warning("No se encontraron datos validos para calcular.")


# ============================================================
# NOMINA COMPLETA (IRT) — Manual con base IMSS libre
# ============================================================
elif tipo == "nomina":
    st.markdown('<div class="section-header"><h3>Nomina IRT — Grupos de Empleados</h3></div>', unsafe_allow_html=True)
    st.markdown("Agrega los grupos de empleados con sus puestos, sueldos actuales y base IMSS deseada.")

    if "grupos" not in st.session_state:
        st.session_state.grupos = []

    with st.expander("Agregar grupo de empleados", expanded=len(st.session_state.grupos) == 0):
        col1, col2 = st.columns(2)
        with col1:
            puesto_sel = st.selectbox("Puesto (referencia)", options=LISTA_PUESTOS, key="puesto_nuevo")
            if puesto_sel == "Otro (personalizado)":
                puesto_custom = st.text_input("Nombre del puesto", key="puesto_custom")
            num_empleados = st.number_input("Numero de empleados", min_value=1, value=10, step=1, key="num_emp_nuevo")

        with col2:
            tipo_sueldo_manual = st.radio(
                "El sueldo capturado es:",
                options=["Bruto", "Neto"],
                index=0,
                horizontal=True,
                key="tipo_sueldo_manual"
            )
            sueldo_input = st.number_input("Sueldo mensual ($)", min_value=1000, value=15000, step=500, key="sueldo_nuevo")
            st.markdown(f"**Prima de riesgo:** {prima_riesgo_global:.5f}%")

        # Compute effective bruto for base IMSS comparison
        sueldo_ref = sueldo_input
        if tipo_sueldo_manual == "Neto":
            sueldo_ref = neto_a_bruto(sueldo_input, clase_riesgo_global, prima_riesgo_global)
            st.info(f"Neto capturado: **{fmt_moneda(sueldo_input)}** → Bruto estimado: **{fmt_moneda(sueldo_ref)}**")

        # Base IMSS libre
        if puesto_sel != "Otro (personalizado)":
            min_prof_default = PUESTOS_PROFESIONALES.get(puesto_sel, SALARIO_MINIMO_MENSUAL)
        else:
            min_prof_default = SALARIO_MINIMO_MENSUAL

        st.markdown("---")
        st.markdown(f"**Base IMSS mensual** — Minimo profesional de referencia: **{fmt_moneda(min_prof_default)}**")

        # Base IMSS inteligente según nivel de esquema
        base_sugerida = calcular_base_imss_inteligente(sueldo_ref, min_prof_default, nivel_esquema)

        opciones_base_keys = [
            f"Sugerida nivel {nivel_esquema} ({fmt_moneda(base_sugerida)})",
            f"Salario minimo ({fmt_moneda(SALARIO_MINIMO_MENSUAL)})",
            f"Minimo profesional ({fmt_moneda(min_prof_default)})",
            "Personalizada",
        ]
        opciones_base_vals = [base_sugerida, SALARIO_MINIMO_MENSUAL, min_prof_default, None]

        opcion_base_sel = st.radio("Base IMSS sugerida:", options=opciones_base_keys, index=0, horizontal=True, key="opcion_base")
        base_valor = opciones_base_vals[opciones_base_keys.index(opcion_base_sel)]

        if base_valor is None:
            base_imss_libre = st.number_input(
                "Base IMSS mensual ($)",
                min_value=float(SALARIO_MINIMO_MENSUAL),
                value=float(min_prof_default),
                step=500.0,
                key="base_imss_libre",
                help="Puedes poner cualquier monto >= salario minimo."
            )
        else:
            base_imss_libre = base_valor
            st.markdown(f"Base IMSS seleccionada: **{fmt_moneda(base_imss_libre)}**")

        # Warning if no IRT differential
        if base_imss_libre >= sueldo_ref:
            st.warning(f"No hay diferencial IRT: la base IMSS ({fmt_moneda(base_imss_libre)}) es >= sueldo bruto ({fmt_moneda(sueldo_ref)}). No conviene migrar a IRT. Reduce la base IMSS o considera otro esquema.")

        if st.button("Agregar grupo", type="primary", use_container_width=True):
            puesto_nombre = puesto_custom if puesto_sel == "Otro (personalizado)" else puesto_sel

            # Convert neto to bruto if needed
            if tipo_sueldo_manual == "Neto":
                sueldo_bruto_final = neto_a_bruto(sueldo_input, clase_riesgo_global, prima_riesgo_global)
            else:
                sueldo_bruto_final = sueldo_input

            grupo_data = {
                "puesto": puesto_nombre,
                "num_empleados": num_empleados,
                "sueldo_bruto": sueldo_bruto_final,
                "clase_riesgo": clase_riesgo_global,
                "minimo_profesional": base_imss_libre,
                "prima_riesgo": prima_riesgo_global,
                "es_neto": tipo_sueldo_manual == "Neto",
            }
            if tipo_sueldo_manual == "Neto":
                grupo_data["sueldo_neto_original"] = sueldo_input
            st.session_state.grupos.append(grupo_data)
            st.rerun()

    if st.session_state.grupos:
        st.markdown('<div class="section-header"><h3>Grupos capturados</h3></div>', unsafe_allow_html=True)
        rows_display = []
        for g in st.session_state.grupos:
            row_d = {
                "Puesto": g["puesto"],
                "# Empleados": g["num_empleados"],
                "Base IMSS": fmt_moneda(g["minimo_profesional"]),
            }
            if g.get("es_neto"):
                row_d["Sueldo Neto"] = fmt_moneda(g["sueldo_neto_original"])
                row_d["Bruto Equiv."] = fmt_moneda(g["sueldo_bruto"])
            else:
                row_d["Sueldo Bruto"] = fmt_moneda(g["sueldo_bruto"])
            rows_display.append(row_d)
        df_display = pd.DataFrame(rows_display)
        st.dataframe(df_display, use_container_width=True, hide_index=True)

        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            if st.button("Eliminar ultimo grupo"):
                st.session_state.grupos.pop()
                st.rerun()
        with col2:
            if st.button("Limpiar todos"):
                st.session_state.grupos = []
                st.rerun()

        st.markdown("---")

        if st.button("CALCULAR PROPUESTA", type="primary", use_container_width=True):
            with st.spinner("Calculando propuesta..."):
                resultados_grupos = []
                any_neto = False
                for g in st.session_state.grupos:
                    r = calcular_grupo_nomina(
                        puesto=g["puesto"],
                        num_empleados=g["num_empleados"],
                        sueldo_bruto=g["sueldo_bruto"],
                        clase_riesgo=g.get("clase_riesgo", clase_riesgo_global),
                        minimo_profesional=g["minimo_profesional"],
                        comision_pct=comision_pct,
                        prima_riesgo=g.get("prima_riesgo", prima_riesgo_global),
                    )
                    if g.get("es_neto"):
                        r["sueldo_neto_original"] = g["sueldo_neto_original"]
                        any_neto = True
                    resultados_grupos.append(r)

                mostrar_resultados_nomina(resultados_grupos, comision_pct, nombre_empresa, contacto, es_neto=any_neto)
    else:
        st.info("Agrega al menos un grupo de empleados para generar la propuesta.")


# ============================================================
# SOLO EXCEDENTES
# ============================================================
elif tipo == "excedentes":
    st.markdown('<div class="section-header"><h3>Administracion de Excedentes</h3></div>', unsafe_allow_html=True)
    st.markdown("Para bonos, comisiones, viaticos y otros conceptos excedentes al salario base.")

    monto_excedente = st.number_input("Monto neto de excedentes a dispersar ($)",
                                       min_value=1000, value=100000, step=5000,
                                       help="Monto que el trabajador recibira como excedente (bono, comision, viatico, etc.)")

    tiene_esquema_formal = st.radio(
        "El cliente tiene esquema formal para estos pagos?",
        ["Si — comparar vs nomina", "No — solo facturacion"],
        index=0,
        horizontal=True,
        key="esquema_formal_exc",
    )
    mostrar_comparativo_exc = tiene_esquema_formal.startswith("Si")

    if st.button("CALCULAR PROPUESTA", type="primary", use_container_width=True):
        r = calcular_excedentes(monto_excedente, comision_pct)

        costo_exc_pre_iva = r['monto_excedente'] + r['comision']
        pct_ahorro_exc = (r['ahorro_mensual'] / r['costo_hipotetico_nomina'] * 100) if r['costo_hipotetico_nomina'] > 0 else 0

        if mostrar_comparativo_exc:
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <h3>Costo Empresa (pre-IVA)</h3>
                    <p>{fmt_moneda(costo_exc_pre_iva)}</p>
                    <p class="sub">Excedente + Comision</p>
                </div>""", unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="ahorro-verde">
                    <h3>Ahorro Mensual</h3>
                    <p>{fmt_moneda(r['ahorro_mensual'])}</p>
                    <p class="sub">{pct_ahorro_exc:.1f}% vs nomina</p>
                </div>""", unsafe_allow_html=True)
            with col3:
                st.markdown(f"""
                <div class="ahorro-verde">
                    <h3>Ahorro Anual</h3>
                    <p>{fmt_moneda(r['ahorro_anual'])}</p>
                    <p class="sub">Proyeccion 12 meses</p>
                </div>""", unsafe_allow_html=True)

            # IVA acreditable + nota deducibilidad
            st.markdown(f"**IVA (acreditable):** {fmt_moneda(r['iva'])} — no es costo para la empresa.")
            st.caption("La comision es 100% deducible para ISR empresarial y el IVA es acreditable.")

            # Ahorro validation
            if r['ahorro_mensual'] <= 0:
                st.warning("El ahorro es negativo. Recomendacion: revisar el monto o considerar otro esquema.")
            elif pct_ahorro_exc < 3:
                st.warning(f"El ahorro es minimo ({pct_ahorro_exc:.1f}%). Considerar alternativas.")

            # Comparativo visual
            import altair as alt
            st.markdown('<div class="section-header"><h3>Comparativo Excedentes vs Nomina</h3></div>', unsafe_allow_html=True)
            chart_data = pd.DataFrame({
                "Esquema": ["Costo Nomina Hipotetico", "Costo Excedentes (pre-IVA)"],
                "Monto": [r["costo_hipotetico_nomina"], costo_exc_pre_iva]
            })
            chart = alt.Chart(chart_data).mark_bar(cornerRadiusTopLeft=6, cornerRadiusTopRight=6).encode(
                x=alt.X("Esquema:N", sort=None, axis=alt.Axis(labelAngle=0)),
                y=alt.Y("Monto:Q", axis=alt.Axis(format="$,.0f")),
                color=alt.Color("Esquema:N", scale=alt.Scale(
                    domain=["Costo Nomina Hipotetico", "Costo Excedentes (pre-IVA)"],
                    range=["#1B3A5C", "#27AE60"]
                ), legend=None),
                tooltip=[alt.Tooltip("Esquema:N"), alt.Tooltip("Monto:Q", format="$,.2f")]
            ).properties(height=300)
            st.altair_chart(chart, use_container_width=True)

            # Comparison table for Excedentes
            st.markdown('<div class="section-header"><h3>Comparativo: Nomina vs Excedentes</h3></div>', unsafe_allow_html=True)
            comp_html = '<table class="comp-table"><tr><th>Concepto</th><th class="num">Si pagara por Nomina</th><th class="num">Esquema Excedentes</th></tr>'
            comp_html += f'<tr><td>Monto a dispersar</td><td class="num">{fmt_moneda(r["monto_excedente"])}</td><td class="num">{fmt_moneda(r["monto_excedente"])}</td></tr>'
            comp_html += f'<tr><td>ISR gravable (Art. 96)</td><td class="num">{fmt_moneda(r["isr_hipotetico"])}</td><td class="num">{fmt_moneda(0)}</td></tr>'
            comp_html += f'<tr><td>IMSS patronal</td><td class="num">{fmt_moneda(r["imss_pat_hipotetico"])}</td><td class="num">{fmt_moneda(0)}</td></tr>'
            comp_html += f'<tr><td>INFONAVIT</td><td class="num">{fmt_moneda(r["infonavit_hipotetico"])}</td><td class="num">{fmt_moneda(0)}</td></tr>'
            comp_html += f'<tr><td>ISN ({tasa_isn*100:.1f}%)</td><td class="num">{fmt_moneda(r["isn_hipotetico"])}</td><td class="num">{fmt_moneda(0)}</td></tr>'
            comp_html += f'<tr><td>Prestaciones de ley</td><td class="num">{fmt_moneda(r["prestaciones_hipotetico"])}</td><td class="num">{fmt_moneda(0)}</td></tr>'
            comp_html += f'<tr><td>Comision</td><td class="num">{fmt_moneda(0)}</td><td class="num">{fmt_moneda(r["comision"])}</td></tr>'
            comp_html += f'<tr><td>IVA (acreditable)</td><td class="num">—</td><td class="num">{fmt_moneda(r["iva"])}</td></tr>'
            comp_html += f'<tr class="ahorro-row"><td><strong>Costo total empresa</strong></td><td class="num">{fmt_moneda(r["costo_hipotetico_nomina"])}</td><td class="num">{fmt_moneda(r["monto_excedente"] + r["comision"])}</td></tr>'
            comp_html += f'<tr class="ahorro-row"><td><strong>Ahorro mensual (pre-IVA)</strong></td><td class="num" colspan="2" style="text-align:center">{fmt_moneda(r["ahorro_mensual"])}</td></tr>'
            comp_html += '</table>'
            st.markdown(comp_html, unsafe_allow_html=True)

        else:
            # Solo facturación — sin comparativo de ahorro
            st.markdown('<div class="section-header"><h3>Facturacion de Excedentes</h3></div>', unsafe_allow_html=True)
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <h3>Monto a Dispersar</h3>
                    <p>{fmt_moneda(r['monto_excedente'])}</p>
                </div>""", unsafe_allow_html=True)
            with col2:
                st.markdown(f"""
                <div class="metric-card">
                    <h3>Total Factura (IVA incl.)</h3>
                    <p>{fmt_moneda(r['total_factura'])}</p>
                </div>""", unsafe_allow_html=True)

            fact_html = '<table class="comp-table"><tr><th>Concepto</th><th class="num">Monto</th></tr>'
            fact_html += f'<tr><td>Excedente a dispersar</td><td class="num">{fmt_moneda(r["monto_excedente"])}</td></tr>'
            fact_html += f'<tr><td>Comision ({comision_pct}%)</td><td class="num">{fmt_moneda(r["comision"])}</td></tr>'
            fact_html += f'<tr><td>Subtotal</td><td class="num">{fmt_moneda(costo_exc_pre_iva)}</td></tr>'
            fact_html += f'<tr><td>IVA (16%)</td><td class="num">{fmt_moneda(r["iva"])}</td></tr>'
            fact_html += f'<tr class="ahorro-row"><td><strong>Total Factura</strong></td><td class="num">{fmt_moneda(r["total_factura"])}</td></tr>'
            fact_html += '</table>'
            st.markdown(fact_html, unsafe_allow_html=True)
            st.caption("La comision es 100% deducible para ISR empresarial y el IVA es acreditable.")

        # Descargas
        st.markdown("---")
        st.markdown('<div class="section-header"><h3>Descargar Propuesta</h3></div>', unsafe_allow_html=True)
        datos_cliente = {
            "nombre_empresa": nombre_empresa or "Cliente",
            "contacto": contacto or "Sin especificar",
            "comision_pct": comision_pct,
        }
        buffer_word = generar_propuesta_word(datos_cliente, "excedentes", r)

        # Excel for excedentes
        buf_xl = io.BytesIO()
        df_xl = pd.DataFrame({
            "Concepto": ["Excedente a dispersar", "Comision", "Subtotal", "IVA", "Total factura",
                         "Costo si pagara por nomina", "Ahorro mensual", "Ahorro anual"],
            "Monto": [r["monto_excedente"], r["comision"], r["monto_excedente"] + r["comision"],
                      r["iva"], r["total_factura"],
                      r["costo_hipotetico_nomina"], r["ahorro_mensual"], r["ahorro_anual"]]
        })
        df_xl.to_excel(buf_xl, index=False, sheet_name="Excedentes")
        buf_xl.seek(0)

        nombre_base = f"{nombre_empresa or 'Cliente'}_{datetime.now().strftime('%Y%m%d')}"
        col_w, col_e = st.columns(2)
        with col_w:
            st.download_button(
                label="DESCARGAR WORD",
                data=buffer_word,
                file_name=f"Propuesta_Excedentes_{nombre_base}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                type="primary",
                use_container_width=True,
            )
        with col_e:
            st.download_button(
                label="DESCARGAR EXCEL",
                data=buf_xl,
                file_name=f"Propuesta_Excedentes_{nombre_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary",
                use_container_width=True,
            )


# ============================================================
# SOCIEDAD CIVIL — Piramidacion como modo default
# ============================================================
elif tipo == "sc":
    st.markdown('<div class="section-header"><h3>Sociedad Civil — Directivos y Gerenciales</h3></div>', unsafe_allow_html=True)
    st.markdown("Para perfiles que no cotizan en IMSS: directivos, socios, duenos.")

    if "directivos" not in st.session_state:
        st.session_state.directivos = []

    with st.expander("Agregar directivo/socio", expanded=len(st.session_state.directivos) == 0):
        nombre_dir = st.text_input("Nombre o identificador", placeholder="Director General", key="nombre_dir")

        col_sc1, col_sc2 = st.columns(2)
        with col_sc1:
            st.markdown("**Retencion ISR**")
            pct_anticipo = st.radio("% Anticipo por remanente", options=[10, 20], horizontal=True, key="pct_ant")
        with col_sc2:
            st.markdown("**Modo de captura**")
            modo_sc = st.radio(
                "Como quieres capturar?",
                options=["Piramidar (neto → bruto)", "Ingreso bruto directo"],
                index=0,
                key="modo_sc",
                horizontal=True,
            )

        piramidar = modo_sc == "Piramidar (neto → bruto)"

        if piramidar:
            neto_deseado = st.number_input("Neto deseado ($)", min_value=10000, value=100000, step=5000, key="neto_des")
            ingreso_total = 0
        else:
            ingreso_total = st.number_input("Ingreso total mensual ($)", min_value=10000, value=100000, step=5000, key="ing_total")
            neto_deseado = 0

        if st.button("Agregar directivo", type="primary", use_container_width=True):
            st.session_state.directivos.append({
                "nombre": nombre_dir or f"Directivo {len(st.session_state.directivos)+1}",
                "ingreso_total": ingreso_total,
                "pct_anticipo": pct_anticipo,
                "piramidar": piramidar,
                "neto_deseado": neto_deseado,
            })
            st.rerun()

    if st.session_state.directivos:
        st.markdown('<div class="section-header"><h3>Directivos capturados</h3></div>', unsafe_allow_html=True)
        for i, d in enumerate(st.session_state.directivos):
            modo = f"Piramidar → Neto {fmt_moneda(d['neto_deseado'])}" if d["piramidar"] else f"Ingreso {fmt_moneda(d['ingreso_total'])}"
            st.write(f"**{i+1}. {d['nombre']}** — {modo} — Anticipo {d['pct_anticipo']}%")

        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button("Eliminar ultimo"):
                st.session_state.directivos.pop()
                st.rerun()

        st.markdown("---")

        if st.button("CALCULAR PROPUESTA", type="primary", use_container_width=True):
            resultados_sc = []
            for d in st.session_state.directivos:
                r = calcular_sociedad_civil(
                    ingreso_total=d["ingreso_total"],
                    pct_anticipo=d["pct_anticipo"],
                    comision_pct=comision_pct,
                    piramidar=d["piramidar"],
                    neto_deseado=d["neto_deseado"],
                )
                r["nombre"] = d["nombre"]
                resultados_sc.append(r)

            for r in resultados_sc:
                st.markdown(f'<div class="section-header"><h3>{r["nombre"]}</h3></div>', unsafe_allow_html=True)

                costo_sc_pre_iva = r['ingreso_total'] + r['comision']

                # Clear cost summary when piramidando
                if r.get("piramidar"):
                    st.markdown(
                        f"Para que el directivo reciba **{fmt_moneda(r['neto_total'])}** netos, "
                        f"el ingreso bruto necesario es **{fmt_moneda(r['ingreso_total'])}** "
                        f"y el costo empresa (pre-IVA) es **{fmt_moneda(costo_sc_pre_iva)}**."
                    )

                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h3>Costo Empresa (pre-IVA)</h3>
                        <p>{fmt_moneda(costo_sc_pre_iva)}</p>
                        <p class="sub">Ingreso + Comision</p>
                    </div>""", unsafe_allow_html=True)
                with col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h3>Neto Directivo</h3>
                        <p>{fmt_moneda(r['neto_total'])}</p>
                        <p class="sub">Anticipo + Renta</p>
                    </div>""", unsafe_allow_html=True)
                with col3:
                    pct_ahorro_sc = (r['ahorro_cliente_mensual'] / r['costo_nomina_100'] * 100) if r['costo_nomina_100'] > 0 else 0
                    st.markdown(f"""
                    <div class="ahorro-verde">
                        <h3>Ahorro Mensual</h3>
                        <p>{fmt_moneda(r['ahorro_cliente_mensual'])}</p>
                        <p class="sub">{pct_ahorro_sc:.1f}% vs nomina</p>
                    </div>""", unsafe_allow_html=True)
                with col4:
                    st.markdown(f"""
                    <div class="ahorro-verde">
                        <h3>Ahorro Anual</h3>
                        <p>{fmt_moneda(r['ahorro_cliente_anual'])}</p>
                        <p class="sub">Proyeccion 12 meses</p>
                    </div>""", unsafe_allow_html=True)

                # IVA acreditable + nota deducibilidad
                st.markdown(f"**IVA (acreditable):** {fmt_moneda(r['iva'])} — no es costo para la empresa.")
                st.caption("La comision es 100% deducible para ISR empresarial y el IVA es acreditable.")

                # Ahorro validation
                if r['ahorro_cliente_mensual'] <= 0:
                    st.warning(f"Ahorro negativo para {r['nombre']}. Recomendacion: revisar % anticipo o considerar otro esquema.")
                elif pct_ahorro_sc < 3:
                    st.warning(f"Ahorro minimo ({pct_ahorro_sc:.1f}%) para {r['nombre']}. Considerar alternativas.")

                st.markdown("")

                # Comparativo visual SC vs Nomina
                import altair as alt
                chart_sc = pd.DataFrame({
                    "Esquema": ["Nomina 100%", "Sociedad Civil"],
                    "Costo": [r["costo_nomina_100"], r["ingreso_total"] + r["comision"]],
                    "Neto": [r["neto_nomina"], r["neto_total"]],
                })
                c_cost = alt.Chart(chart_sc).mark_bar(cornerRadiusTopLeft=6, cornerRadiusTopRight=6).encode(
                    x=alt.X("Esquema:N", sort=None, axis=alt.Axis(labelAngle=0)),
                    y=alt.Y("Costo:Q", axis=alt.Axis(format="$,.0f"), title="Costo Empresa (pre-IVA)"),
                    color=alt.Color("Esquema:N", scale=alt.Scale(
                        domain=["Nomina 100%", "Sociedad Civil"],
                        range=["#1B3A5C", "#27AE60"]
                    ), legend=None),
                    tooltip=[alt.Tooltip("Esquema:N"), alt.Tooltip("Costo:Q", format="$,.2f")]
                ).properties(height=280, title="Costo Empresa")

                c_neto = alt.Chart(chart_sc).mark_bar(cornerRadiusTopLeft=6, cornerRadiusTopRight=6).encode(
                    x=alt.X("Esquema:N", sort=None, axis=alt.Axis(labelAngle=0)),
                    y=alt.Y("Neto:Q", axis=alt.Axis(format="$,.0f"), title="Neto Directivo"),
                    color=alt.Color("Esquema:N", scale=alt.Scale(
                        domain=["Nomina 100%", "Sociedad Civil"],
                        range=["#1B3A5C", "#C9A962"]
                    ), legend=None),
                    tooltip=[alt.Tooltip("Esquema:N"), alt.Tooltip("Neto:Q", format="$,.2f")]
                ).properties(height=280, title="Neto Directivo")

                chart_col1, chart_col2 = st.columns(2)
                with chart_col1:
                    st.altair_chart(c_cost, use_container_width=True)
                with chart_col2:
                    st.altair_chart(c_neto, use_container_width=True)

                # P5: Comparison table SC vs Nomina
                comp_html = '<table class="comp-table"><tr><th>Concepto</th><th class="num">Nomina 100%</th><th class="num">Sociedad Civil</th></tr>'
                comp_html += f'<tr><td>Ingreso bruto</td><td class="num">{fmt_moneda(r["ingreso_total"])}</td><td class="num">{fmt_moneda(r["ingreso_total"])}</td></tr>'
                comp_html += f'<tr><td>ISR retenido</td><td class="num">{fmt_moneda(r["isr_nomina"]["isr_neto"])}</td><td class="num">{fmt_moneda(r["isr_anticipo"]["isr_neto"])}</td></tr>'
                comp_html += f'<tr><td>IMSS obrero</td><td class="num">{fmt_moneda(0)}</td><td class="num">{fmt_moneda(0)}</td></tr>'
                comp_html += f'<tr><td>Anticipo por remanente ({r["pct_anticipo"]}%)</td><td class="num">—</td><td class="num">{fmt_moneda(r["anticipo"])}</td></tr>'
                comp_html += f'<tr><td>Renta vitalicia (exenta)</td><td class="num">—</td><td class="num">{fmt_moneda(r["renta"])}</td></tr>'
                comp_html += f'<tr><td><strong>Neto directivo</strong></td><td class="num">{fmt_moneda(r["neto_nomina"])}</td><td class="num">{fmt_moneda(r["neto_total"])}</td></tr>'
                comp_html += f'<tr><td>IMSS patronal + ISN</td><td class="num">{fmt_moneda(r["costo_nomina_100"] - r["ingreso_total"])}</td><td class="num">{fmt_moneda(0)}</td></tr>'
                comp_html += f'<tr><td>Comision SC</td><td class="num">{fmt_moneda(0)}</td><td class="num">{fmt_moneda(r["comision"])}</td></tr>'
                comp_html += f'<tr class="ahorro-row"><td><strong>Costo empresa (pre-IVA)</strong></td><td class="num">{fmt_moneda(r["costo_nomina_100"])}</td><td class="num">{fmt_moneda(r["ingreso_total"] + r["comision"])}</td></tr>'
                comp_html += f'<tr class="ahorro-row"><td><strong>Ahorro mensual</strong></td><td class="num" colspan="2" style="text-align:center">{fmt_moneda(r["ahorro_cliente_mensual"])}</td></tr>'
                comp_html += f'<tr class="ahorro-row"><td><strong>Ahorro anual</strong></td><td class="num" colspan="2" style="text-align:center">{fmt_moneda(r["ahorro_cliente_anual"])}</td></tr>'
                comp_html += '</table>'
                st.markdown(comp_html, unsafe_allow_html=True)
                st.markdown("---")

            # Descargas
            st.markdown('<div class="section-header"><h3>Descargar Propuesta</h3></div>', unsafe_allow_html=True)
            datos_cliente = {
                "nombre_empresa": nombre_empresa or "Cliente",
                "contacto": contacto or "Sin especificar",
                "comision_pct": comision_pct,
            }
            try:
                buffer_word = generar_propuesta_word(datos_cliente, "sc", resultados_sc)
            except Exception as e:
                st.warning(f"No se pudo generar el Word: {e}")
                buffer_word = None

            # Excel for SC
            buf_xl = io.BytesIO()
            rows_xl = []
            for r in resultados_sc:
                rows_xl.append({
                    "Directivo": r["nombre"],
                    "Ingreso Bruto": round(r["ingreso_total"], 2),
                    "Anticipo": round(r["anticipo"], 2),
                    "ISR Anticipo": round(r["isr_anticipo"]["isr_neto"], 2),
                    "Renta Vitalicia": round(r["renta"], 2),
                    "Neto Total": round(r["neto_total"], 2),
                    "Comision": round(r["comision"], 2),
                    "IVA": round(r["iva"], 2),
                    "Total Factura": round(r["total_factura"], 2),
                    "Costo Nomina 100%": round(r["costo_nomina_100"], 2),
                    "Ahorro Mensual": round(r["ahorro_cliente_mensual"], 2),
                    "Ahorro Anual": round(r["ahorro_cliente_anual"], 2),
                })
            pd.DataFrame(rows_xl).to_excel(buf_xl, index=False, sheet_name="Sociedad Civil")
            buf_xl.seek(0)

            nombre_base = f"{nombre_empresa or 'Cliente'}_{datetime.now().strftime('%Y%m%d')}"
            col_w, col_e = st.columns(2)
            with col_w:
                if buffer_word:
                    st.download_button(
                        label="DESCARGAR WORD",
                        data=buffer_word,
                        file_name=f"Propuesta_SC_{nombre_base}.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        type="primary",
                        use_container_width=True,
                    )
            with col_e:
                st.download_button(
                    label="DESCARGAR EXCEL",
                    data=buf_xl,
                    file_name=f"Propuesta_SC_{nombre_base}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="secondary",
                    use_container_width=True,
                )
    else:
        st.info("Agrega al menos un directivo/socio para generar la propuesta.")

# === FOOTER ===
st.markdown("---")
st.markdown(
    f'<p style="text-align: center; color: #A0AEC0; font-size: 0.75rem; margin-top: 1rem;">'
    f'Sistema de Cotizacion Fiscal v3.0 &middot; Ano fiscal 2026 &middot; {estado_sel} (ISN {tasa_isn*100:.2f}%) &middot; '
    f'Uso interno exclusivo &middot; {datetime.now().strftime("%Y")}</p>',
    unsafe_allow_html=True
)
